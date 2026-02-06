"""Excel report: output formatting, styling, price-sources sheet."""
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

from price_search import load_web_price_cache, save_web_price_cache


def format_dataframe_for_output(df, is_watchlist=False):
    """
    Format DataFrame to match template.

    Template columns:
    Type, Asset, WKN, ISIN, Total Quantity, Invested EUR, Profit EUR, Profit %,
    News Sentiment, Recommendation, Reasoning, Change Reasoning
    """
    formatted = pd.DataFrame()
    n_rows = len(df)

    formatted["Type"] = ["Watchlist" if is_watchlist else "Holdings"] * n_rows
    formatted["Asset"] = df["Asset"].values if "Asset" in df.columns else [""] * n_rows
    formatted["WKN"] = df["WKN"].values if "WKN" in df.columns else [""] * n_rows
    formatted["ISIN"] = df["ISIN"].values if "ISIN" in df.columns else [""] * n_rows

    if "Stueckzahl" in df.columns:
        formatted["Total Quantity"] = df["Stueckzahl"].values
    elif "Anzahl" in df.columns:
        formatted["Total Quantity"] = df["Anzahl"].values
    else:
        formatted["Total Quantity"] = [""] * n_rows

    formatted["Invested EUR"] = df["Invest"].values if "Invest" in df.columns else [""] * n_rows

    profit_eur = []
    for idx, row in df.iterrows():
        try:
            invest = float(row.get("Invest", 0) or 0)
            current_price = None
            if row.get("Web_Price") and isinstance(row.get("Web_Price"), dict):
                current_price = row["Web_Price"].get("price")
            elif row.get("Alpaca_Price") and isinstance(row.get("Alpaca_Price"), dict):
                current_price = row["Alpaca_Price"].get("mid_price")
            if current_price and invest > 0:
                qty = float(row.get("Stueckzahl", 0) or row.get("Anzahl", 0) or 0)
                current_value = current_price * qty
                profit = current_value - invest
                profit_eur.append(round(profit, 2))
            else:
                profit_eur.append("")
        except Exception:
            profit_eur.append("")
    formatted["Profit EUR"] = profit_eur

    profit_pct = []
    for idx, row in df.iterrows():
        try:
            invest = float(row.get("Invest", 0) or 0)
            current_price = None
            if row.get("Web_Price") and isinstance(row.get("Web_Price"), dict):
                current_price = row["Web_Price"].get("price")
            elif row.get("Alpaca_Price") and isinstance(row.get("Alpaca_Price"), dict):
                current_price = row["Alpaca_Price"].get("mid_price")
            if current_price and invest > 0:
                qty = float(row.get("Stueckzahl", 0) or row.get("Anzahl", 0) or 0)
                current_value = current_price * qty
                pct = ((current_value - invest) / invest) * 100
                profit_pct.append(f"{round(pct, 1)}%")
            else:
                profit_pct.append("")
        except Exception:
            profit_pct.append("")
    formatted["Profit %"] = profit_pct

    sentiment = []
    for idx, row in df.iterrows():
        news_text = ""
        news_data = row.get("News")
        if news_data and isinstance(news_data, dict):
            news_items = news_data.get("news_items", [])
            if news_items:
                headlines = []
                for item in news_items[:3]:
                    if isinstance(item, dict):
                        title = item.get("title", "")
                        source = item.get("source", "")
                        if title:
                            headlines.append(f"{source}: {title[:50]}")
                if headlines:
                    news_text = "; ".join(headlines)
        sentiment.append(news_text if news_text else "")
    formatted["News Sentiment"] = sentiment

    rec = df["Recommendation"].values if "Recommendation" in df.columns else None
    if rec is None:
        rec = df["Empfehlung"].values if "Empfehlung" in df.columns else [""] * n_rows
    formatted["Recommendation"] = rec if rec is not None else [""] * n_rows

    reasoning = None
    if "Reasoning" in df.columns:
        reasoning = df["Reasoning"].values
    elif "Begründung" in df.columns:
        reasoning = df["Begründung"].values
    elif "Begruendung" in df.columns:
        reasoning = df["Begruendung"].values
    formatted["Reasoning"] = reasoning if reasoning is not None else [""] * n_rows

    qty_reason = df["quantity_reasoning"].values if "quantity_reasoning" in df.columns else None
    if qty_reason is None:
        qty_reason = df["Grund_fuer_Menge"].values if "Grund_fuer_Menge" in df.columns else [""] * n_rows
    formatted["Change Reasoning"] = qty_reason if qty_reason is not None else [""] * n_rows

    return formatted


def save_analysis_excel(output_file, results, watchlist_results, assets):
    """Save analysis Excel with portfolio, watchlist, and price-sources sheets."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if results:
            main_df = pd.DataFrame(results)
            formatted_main = format_dataframe_for_output(main_df, is_watchlist=False)
            formatted_main.to_excel(writer, sheet_name="Portfolio Analyse", index=False)

        if watchlist_results:
            watch_df = pd.DataFrame(watchlist_results)
            formatted_watch = format_dataframe_for_output(watch_df, is_watchlist=True)
            formatted_watch.to_excel(writer, sheet_name="Watchlist Analyse", index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        fill_green_buy = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_green_sell = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        column_widths = {
            "A": 11, "B": 33, "C": 8, "D": 14, "E": 14, "F": 10,
            "G": 10, "H": 10, "I": 23, "J": 27, "K": 50, "L": 50,
        }

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            idx_recommendation = -1
            for cell in worksheet[1]:
                header_val = str(cell.value or "").lower()
                if "recommendation" in header_val or "empfehlung" in header_val:
                    idx_recommendation = cell.column
                    break

            for row_idx in range(2, worksheet.max_row + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = thin_border
                    if isinstance(cell.value, (datetime, pd.Timestamp)):
                        cell.number_format = "DD.MM.YYYY HH:MM:SS"

                if idx_recommendation != -1:
                    cell_rec = worksheet.cell(row=row_idx, column=idx_recommendation)
                    val = str(cell_rec.value or "").lower()
                    if "sell" in val or "verkauf" in val:
                        if "partial" not in val and "teil" not in val:
                            cell_rec.fill = fill_red
                        else:
                            cell_rec.fill = fill_yellow
                    elif "buy" in val or "kauf" in val or "add" in val or "aufstock" in val:
                        cell_rec.fill = fill_green_buy
                    elif "hold" in val or "halten" in val:
                        cell_rec.fill = fill_green_sell
                    elif "reduce" in val or "reduzier" in val:
                        cell_rec.fill = fill_yellow

            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            for col in worksheet.columns:
                col_letter = col[0].column_letter
                if col_letter not in column_widths:
                    max_length = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    if adjusted_width < 10:
                        adjusted_width = 12
                    worksheet.column_dimensions[col_letter].width = adjusted_width

        cache = load_web_price_cache()
        price_sources = []
        for i, result in enumerate(results):
            asset = assets[i]
            asset_name = result.get("Asset", "")
            ticker = result.get("Ticker", "")
            cache_key = ticker or asset.get("ISIN", "")
            source_type = ""
            source_detail = ""
            price_value = None

            if result.get("Alpaca_Price") and "mid_price" in result.get("Alpaca_Price", {}):
                source_type = "Alpaca"
                source_detail = "API"
                price_value = result["Alpaca_Price"].get("mid_price")
            elif result.get("Web_Price"):
                web_data = result["Web_Price"]
                if "source" in web_data:
                    source_type = web_data.get("source", "Web")
                    source_detail = web_data.get("url", "")
                    price_value = web_data.get("price")
                    if cache_key and source_detail:
                        cache[cache_key] = {"source": source_type, "url": source_detail}

            if cache_key in cache and not source_type:
                source_type = cache[cache_key].get("source", "Cached")
                source_detail = cache[cache_key].get("url", "")

            price_sources.append({
                "Asset": asset_name,
                "Ticker/ISIN": ticker,
                "Source": source_type or "None",
                "Source Detail / URL": source_detail,
                "Price": price_value,
                "Fetched At": result.get("FetchedAt", ""),
            })

        save_web_price_cache(cache)
        if price_sources:
            price_df = pd.DataFrame(price_sources)
            price_df.to_excel(writer, sheet_name="Price Sources", index=False)
