import sys
import os
import json
import subprocess
import asyncio
import pandas as pd
import time
from datetime import datetime
import traceback
import requests
import xml.etree.ElementTree as ET
import anthropic
import signal
import re
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# --- PATH CONFIGURATION & IMPORTS ---
if sys.platform == "win32":
    MCP_BASE = r"C:\Users\HMz\Documents\Source\McpServer"
    sys.path.append(os.path.join(MCP_BASE, "tiingo-mcp-server"))
    try:
        from tiingo_mcp_server.tiingo_functions import (
            get_ticker_price_sync as tiingo_get_price,
            get_news_sync as tiingo_get_news,
        )
    except ImportError as e:
        print(f"❌ Critical Import Error (Tiingo): {e}")
        sys.exit(1)
else:
    # Linux: Tiingo runs via subprocess in its own venv (different Python versions OK)
    MCP_BASE = "/home/dev/proj/quantrosoft/mcp-server"
    TIINGO_MCP_PATH = os.path.join(MCP_BASE, "tiingo-mcp-server")
    TIINGO_PYTHON = os.path.join(TIINGO_MCP_PATH, ".venv", "bin", "python")
    TIINGO_RUNNER = os.path.join(TIINGO_MCP_PATH, "tiingo_runner.py")

    def _run_tiingo(cmd: list) -> list | dict:
        """Run tiingo_runner.py via tiingo venv; returns parsed JSON or [] on error."""
        if not os.path.isfile(TIINGO_PYTHON):
            raise RuntimeError(f"Tiingo venv not found at {TIINGO_PYTHON}. Install: python -m venv .venv && pip install -r requirements.txt")
        if not os.path.isfile(TIINGO_RUNNER):
            raise RuntimeError(f"Tiingo runner not found at {TIINGO_RUNNER}")
        env = os.environ.copy()
        result = subprocess.run(
            [TIINGO_PYTHON, TIINGO_RUNNER] + cmd,
            capture_output=True,
            text=True,
            cwd=TIINGO_MCP_PATH,
            env=env,
            timeout=60,
        )
        if result.returncode != 0:
            err = result.stderr or result.stdout or "Unknown error"
            raise RuntimeError(f"Tiingo runner failed: {err}")
        out = result.stdout.strip()
        if not out:
            return []
        data = json.loads(out)
        if isinstance(data, dict) and "error" in data:
            raise RuntimeError(data["error"])
        return data

    def tiingo_get_price(ticker: str):
        """Get ticker price via tiingo venv subprocess."""
        return _run_tiingo(["get_ticker_price", ticker])

    def tiingo_get_news(tickers=None, limit: int = 20, **kwargs):
        """Get news via tiingo venv subprocess."""
        cmd = ["get_news", "--limit", str(limit)]
        if tickers is not None:
            t = tickers if isinstance(tickers, str) else ",".join(tickers) if tickers else ""
            if t:
                cmd.extend(["--tickers", t])
        return _run_tiingo(cmd)

    # t_news = tiingo_get_news(tickers=None, limit=50) 

# Alpaca
try:
    from alpaca.data.historical import StockHistoricalDataClient, NewsClient
    from alpaca.data.requests import StockLatestQuoteRequest, NewsRequest
    ALPACA_AVAILABLE = True
    print("Alpaca: OK")
except ImportError as e:
    print("Alpaca import failed:", type(e).__name__, e)
    ALPACA_AVAILABLE = False

# Trade Republic (DISABLED due to 2FA requirement)
TR_AVAILABLE = False
# try:
#     sys.path.append(os.path.join(MCP_BASE, "trade-republic-mcp-server"))
#     from trade_republic_mcp_server.trade_republic_functions import (
#         get_quote as tr_get_quote_async,
#         get_quote_sync as tr_get_quote_sync,
#         search_instruments_sync as tr_search_sync
#     )
#     TR_AVAILABLE = True
#     # Silence TR loggers
#     import logging
#     logging.getLogger('trade_republic_mcp_server').setLevel(logging.CRITICAL)
#     logging.getLogger('pytr').setLevel(logging.CRITICAL)
# except ImportError:
#     TR_AVAILABLE = False

# ReportLab
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

# --- UTILS ---
class Tee(object):
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()
    def flush(self):
        for f in self.files:
            f.flush()

# --- CONFIGURATION ---
Global_EURUSD = None # Must be fetched!

TIINGO_KEY = None
ALPACA_KEY = None
ALPACA_SECRET = None
TR_PHONE = None
TR_PIN = None

def load_env_keys():
    """Load API keys from local env.txt and MCP server dist folders."""
    global TIINGO_KEY, ANTHROPIC_CLIENT, ALPACA_KEY, ALPACA_SECRET, TR_PHONE, TR_PIN
    
    env_paths = [
        "env.txt",  # Local
        os.path.join(MCP_BASE, "tiingo-mcp-server", "dist", "env.txt"),
        os.path.join(MCP_BASE, "alpaca-mcp-server", "dist", "env.txt"),
        os.path.join(MCP_BASE, "ibapi-mcp-server", "dist", "env.txt"),
        os.path.join(MCP_BASE, "trade-republic-mcp-server", "dist", "env.txt"),
    ]
    
    loaded_any = False
    for path in env_paths:
        if os.path.exists(path):
            try:
                with open(path, "r", encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if "=" in line and not line.startswith("#"):
                            key, val = line.split("=", 1)
                            os.environ[key.strip()] = val.strip()
                print(f"✅ Loaded environment from: {path}")
                loaded_any = True
            except Exception as e:
                print(f"⚠️ Error loading keys from {path}: {e}")
    
    if loaded_any:
        TIINGO_KEY = os.environ.get("TIINGO_API_KEY")
        ALPACA_KEY = os.environ.get("APCA_API_KEY_ID") or os.environ.get("ALPACA_API_KEY")
        ALPACA_SECRET = os.environ.get("APCA_API_SECRET_KEY") or os.environ.get("ALPACA_SECRET_KEY")
        TR_PHONE = os.environ.get("TRADE_REPUBLIC_PHONE")
        TR_PIN = os.environ.get("TRADE_REPUBLIC_PIN")
        
        # Priority for Anthropic Key
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if api_key:
            ANTHROPIC_CLIENT = anthropic.AsyncAnthropic(api_key=api_key)
            print(f"✅ Anthropic Client initialized.")
        else:
            print("⚠️ Warning: ANTHROPIC_API_KEY not found in any env.txt.")
    else:
        print("⚠️ Warning: No env.txt files found in any search locations.")

# --- PDF GENERATION ---
def create_pdf(filename, item):
    doc = SimpleDocTemplate(filename, pagesize=letter,
                           topMargin=0.4*inch, bottomMargin=0.4*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Compact styles
    title_style = ParagraphStyle('CompactTitle', parent=styles['Heading1'], 
                                 fontSize=14, spaceAfter=6)
    normal_style = ParagraphStyle('CompactNormal', parent=styles['Normal'],
                                  fontSize=9, leading=11)
    header_style = ParagraphStyle('CompactHeader', parent=styles['Heading2'],
                                  fontSize=11, spaceAfter=4, spaceBefore=6)
    
    asset_name = item.get('Asset') or item.get('Ticker') or "Unknown Asset"
    story.append(Paragraph(f"Analyse: {asset_name}", title_style))
    story.append(Paragraph(f"Datum: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}", normal_style))
    story.append(Spacer(1, 6))

    if 'Ticker' in item:
        story.append(Paragraph(f"<b>Ticker:</b> {item['Ticker']}", normal_style))
        
    if 'Empfehlung' in item:
        rec_color = "black"
        rec_text = str(item['Empfehlung']).lower()
        if 'kauf' in rec_text or 'buy' in rec_text: rec_color = "green"
        elif 'verkauf' in rec_text or 'sell' in rec_text: rec_color = "red"
        
        # Add Quantity if available
        qty_text = ""
        if 'Empfohlene_Stueckzahl' in item and item['Empfohlene_Stueckzahl'] is not None:
             qty_text = f" ({item['Empfohlene_Stueckzahl']} Stk.)"
        
        rec_para = Paragraph(f"<b>Empfehlung:</b> <font color='{rec_color}'>{item['Empfehlung']}{qty_text}</font>", normal_style)
        story.append(rec_para)
        
        # Add Reasoning for Quantity if available
        if 'Grund_fuer_Menge' in item and item['Grund_fuer_Menge']:
             story.append(Paragraph(f"<i>Strategie: {item['Grund_fuer_Menge']}</i>", normal_style))
    
    story.append(Spacer(1, 6))

    # Begründung / News
    if 'Begründung' in item:
        story.append(Paragraph("KI-Analyse & News-Zusammenfassung", header_style))
        text = str(item['Begründung']).replace('\n', '<br/>')
        story.append(Paragraph(text, normal_style))
        story.append(Spacer(1, 6))

    # Raw News Section removed to avoid duplication (AI already lists headlines)

    story.append(Paragraph("Daten", header_style))
    data = []
    # Explicitly filter fields we want to avoid showing raw nested dicts or clutter
    exclude_fields = [
        'Begründung', 'Empfehlung', 'Asset', 'Ticker', 'News',
        'RawNews', 'Tiingo_Price', 'Alpaca_Price', 'Web_Price'
    ]

    for k, v in item.items():
        if k not in exclude_fields and not isinstance(v, (dict, list)):
            val_str = str(v)
            if len(val_str) > 60: val_str = val_str[:60] + "..."
            data.append([k, val_str])

    # Add Price summary if available (clean version)
    current_price = None
    for key in ['Tiingo_Price', 'Alpaca_Price', 'Web_Price']:
        if key in item and 'mid_price' in item[key]:
            current_price = item[key]['mid_price']
            break
        elif key in item and 'price' in item[key]:
            current_price = item[key]['price']
            break
            
    if current_price:
        data.append(['Aktueller Preis', f"{current_price:.4f}"])

    if data:
        t = Table(data, colWidths=[1.5*inch, 4*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(t)

    doc.build(story)

# --- DATA PIPELINE ---

async def get_forex_rate(base="EUR", quote="USD"):
    """
    Fetch current forex rate (e.g. EUR to USD) robustly.
    1. Tiingo API
    2. Web Scraping (Google Finance)
    """
    # 1. Tiingo API (Primary)
    print("    💱 Fetching Forex via Tiingo...")
    if TIINGO_KEY:
        try:
             url = f"https://api.tiingo.com/tiingo/fx/top?tickers={base.lower()}{quote.lower()}&token={TIINGO_KEY}"
             headers = {'Content-Type': 'application/json'}
             resp = requests.get(url, headers=headers, timeout=5)
             if resp.status_code == 200 and resp.json():
                 data = resp.json()[0]
                 rate = data.get('midPrice')
                 if rate:
                     print(f"    💱 Forex via Tiingo: {rate}")
                     return float(rate)
        except Exception as e:
            print(f"    ⚠️ Tiingo Forex failed: {e}")

    # 2. Web Scraping Fallback (Google Finance - simpler structure)
    print("    ⚠️ Tiingo missing, trying Google Finance...")
    try:
        # Google Finance URL
        url = f"https://www.google.com/finance/quote/{base}-{quote}"
        headers = {'User-Agent': 'Mozilla/5.0'}
        resp = requests.get(url, headers=headers, timeout=5)
        # Look for the large text class "YMlKec fxKbKc" usually containing the price
        # Or just regex for the price near the currency code
        match = re.search(r'data-last-price="([\d\.]+)"', resp.text)
        if match:
             rate = float(match.group(1))
             print(f"    💱 Forex via Google Finance: {rate}")
             return rate
             
        # Fallback regex
        match = re.search(r'<div class="YMlKec fxKbKc">([\d\.]+)</div>', resp.text)
        if match:
              rate = float(match.group(1))
              print(f"    💱 Forex via Google Finance: {rate}")
              return rate

    except Exception as e:
        print(f"    ⚠️ Google Finance Forex failed: {e}")
        
    return None

# Web Price Cache (persistent)
if sys.platform == "win32":
    WEB_PRICE_CACHE_FILE = r"G:\Meine Ablage\ShareFile\NewsTrader\web_price_cache.json"
else:
    WEB_PRICE_CACHE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    WEB_PRICE_CACHE_FILE = os.path.join(WEB_PRICE_CACHE_DIR, "web_price_cache.json")

def load_web_price_cache():
    """Load web price source cache from disk."""
    try:
        if os.path.exists(WEB_PRICE_CACHE_FILE):
            with open(WEB_PRICE_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except:
        pass
    return {}

def save_web_price_cache(cache):
    """Save web price source cache to disk."""
    try:
        cache_dir = os.path.dirname(WEB_PRICE_CACHE_FILE)
        if cache_dir and not os.path.exists(cache_dir):
            os.makedirs(cache_dir, exist_ok=True)
        with open(WEB_PRICE_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"    ⚠️ Cache save failed: {e}")

async def deep_dive_price_search(asset, ticker):
    """
    Simulated 'DeepDive' - Searches for asset price on financial sites.
    Uses ISIN, WKN, or ticker as search term.
    """
    isin = asset.get('ISIN')
    wkn = asset.get('WKN')

    # Determine best search identifier (prefer WKN for German derivatives)
    search_id = None
    if wkn and str(wkn) != 'nan' and len(str(wkn)) >= 5:
        search_id = str(wkn)
        print(f"    🤿 DeepDive Search using WKN: {search_id}")
    elif isin and str(isin) != 'nan':
        search_id = str(isin)
        print(f"    🤿 DeepDive Search using ISIN: {search_id}")
    else:
        return {"error": "No ISIN or WKN"}

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'de,en-US;q=0.7,en;q=0.3'
    }

    # 0. QUICK CHECK: Direct site lookups with WKN/ISIN (fastest)
    quick_urls = [
        (f"https://www.onvista.de/suche/{search_id}", "Onvista"),
        (f"https://www.ariva.de/{search_id}", "Ariva"),
        (f"https://www.comdirect.de/inf/search/all.html?SEARCH_VALUE={search_id}", "Comdirect"),
    ]

    for url, source in quick_urls:
        try:
            resp = await asyncio.to_thread(requests.get, url, headers=headers, timeout=8, allow_redirects=True)
            if resp.status_code == 200:
                html = resp.text
                # Look for Brief (Ask) price first
                brief = re.search(r'Brief.*?(\d+,\d{2,4})', html, re.IGNORECASE | re.DOTALL)
                if brief:
                    price = float(brief.group(1).replace(',', '.'))
                    print(f"    ✅ {source} (Brief): {price}")
                    return {"price": price, "source": f"{source} (Brief)", "url": resp.url}

                # Fallback to Geld (Bid)
                geld = re.search(r'Geld.*?(\d+,\d{2,4})', html, re.IGNORECASE | re.DOTALL)
                if geld:
                    price = float(geld.group(1).replace(',', '.'))
                    print(f"    ✅ {source} (Geld): {price}")
                    return {"price": price, "source": f"{source} (Geld)", "url": resp.url}
        except Exception as e:
            pass

    # Keep ISIN for legacy code below
    isin = search_id
    
    # 1. Google Search to find valid deep link
    search_urls = []
    try:
        # Search query: ISIN + Keywords
        query = f"{isin} Kurs aktuell onvista finanzen"
        url = f"https://www.google.com/search?q={query}&num=5" # fetch top 5
        resp = await asyncio.to_thread(requests.get, url, headers=headers, timeout=5)
        
        # Extract Links (Generic regex for hrefs)
        # Matches typical google results: /url?q=https://...
        raw_links = re.findall(r'/url\?q=(https://www\.(?:onvista|finanzen)\.de/[^&]+)', resp.text)
        
        # Also clean links if Google served cleaner HTML (unlikely but possible) 
        # or simple href="https://..."
        if not raw_links:
             raw_links = re.findall(r'href="(https://www\.(?:onvista|finanzen)\.de/[^"]+)"', resp.text)

        for l in raw_links:
            if "google" in l: continue
            search_urls.append(l)
            
    except Exception as e:
        print(f"      ⚠️ Search failed: {e}")

    # Fallback to direct URL construction if search failed
    # Note: Onvista search result often contains the price directly or redirects
    potential_urls = [
         f"https://www.onvista.de/suche/{isin}",
         f"https://www.finanzen.net/suchergebnis?_search={isin}",
    ]
    
    # BNP Direct Check (if ISIN looks like BNP or explicitly requested)
    if "DE000P" in isin or "BNP" in asset.get("Asset", ""):
         bnp_url = f"https://derivate.bnpparibas.com/product-details/{isin}/"
         if bnp_url not in search_urls: potential_urls.append(bnp_url)

    # Merge found search links with potential direct links
    # Prioritize found links, then fallbacks
    urls_to_check = search_urls + [u for u in potential_urls if u not in search_urls]

    # 2. Visit Top Links
    for link in urls_to_check[:5]: # Check top 5
        print(f"    ➡️ Inspecting: {link}")
        try:
            page_resp = await asyncio.to_thread(requests.get, link, headers=headers, timeout=5)
            if page_resp.status_code != 200: continue
            
            html = page_resp.text
            
            # --- 2a. Onvista Price Handling ---
            if "onvista.de" in link:
                # 1. Check for "Brief" (Ask) or "Geld" (Bid) in text
                # HTML often: <span ...>Brief</span><span ...>4,57</span>
                # Or text: "Brief 4,57"
                
                # Matches "Brief" followed within 50 chars by a number
                # Handling whitespace and tags in between
                match = re.search(r'Brief.*?(\d+,\d{2})', html, re.IGNORECASE | re.DOTALL)
                if match:
                    return {"price": float(match.group(1).replace(',', '.')), "source": "Onvista (Brief)", "url": link}
                
                match = re.search(r'Geld.*?(\d+,\d{2})', html, re.IGNORECASE | re.DOTALL)
                if match:
                    # Use Geld if Brief not found
                    return {"price": float(match.group(1).replace(',', '.')), "source": "Onvista (Geld)", "url": link}

                # 2. Meta tag: <meta itemprop="price" content="123.45">
                match = re.search(r'itemprop="price"[^>]*content="([\d\.]+)"', html)
                if match:
                    return {"price": float(match.group(1)), "source": "Onvista (Meta)", "url": link}
                
                # 3. Data value: <data class="price" value="123.45">
                match = re.search(r'value="([\d\.]+)"[^>]*class="[^"]*price[^"]*"', html)
                if match:
                    return {"price": float(match.group(1)), "source": "Onvista (Data)", "url": link}

            # --- 2b. Finanzen.net Price Handling ---
            elif "finanzen.net" in link:
                 # Meta: <meta itemprop="price" content="8.45" />
                 match = re.search(r'itemprop="price"[^>]*content="([\d\.]+)"', html)
                 if match:
                     return {"price": float(match.group(1)), "source": "Finanzen.net", "url": link}
                 
                 # Col Header logic (for search lists)
                 # Look for "Aktuell" column
                 match = re.search(r'col-price"[^>]*>([\d,]+)', html)
                 if match:
                      return {"price": float(match.group(1).replace(',', '.')), "source": "Finanzen.net (Table)", "url": link}

            # --- 2c. BNP Paribas Handling ---
            if "bnpparibas" in link:
                 # 1. Meta Tags (Standard)
                 match = re.search(r'itemprop="price"[^>]*content="([\d\.]+)"', html)
                 if match: return {"price": float(match.group(1)), "source": "BNP (Meta)", "url": link}

                 # 2. JSON Data
                 match = re.search(r'"(?:ask|offer|price|kaufen)"\s*[:=]\s*"?([\d\.]+)"?', html, re.IGNORECASE)
                 if match:
                      return {"price": float(match.group(1)), "source": "BNP (JSON)", "url": link}
                 
                 # 3. Visual Text Scraping
                 # Pattern: "Kaufen" ... "€ X,XX" or "X,XX €"
                 match = re.search(r'Kaufen.*?(\d+,\d{2})', html, re.DOTALL | re.IGNORECASE)
                 if match:
                      return {"price": float(match.group(1).replace(',', '.')), "source": "BNP (Text Kaufen)", "url": link}
                 
                 match = re.search(r'Brief.*?(\d+,\d{2})', html, re.DOTALL | re.IGNORECASE)
                 if match:
                      return {"price": float(match.group(1).replace(',', '.')), "source": "BNP (Text Brief)", "url": link}

        except Exception as e:
            pass

    # 3. FALLBACK: Generic Google Search for price
    print(f"    🔍 Google Search Fallback for price...")
    asset_name = asset.get('Asset', '')

    # Try multiple search queries
    search_queries = [
        f"{isin} Kurs EUR",
        f"{isin} aktueller Preis",
        f'"{asset_name}" Kurs aktuell',
    ]

    for search_query in search_queries:
        try:
            url = f"https://www.google.com/search?q={search_query}"
            resp = await asyncio.to_thread(requests.get, url, headers=headers, timeout=5)
            html = resp.text

            # Pattern 1: Google Finance card - "XX,XX €" or "XX.XX EUR"
            match = re.search(r'data-last-price="([\d\.]+)"', html)
            if match:
                price = float(match.group(1))
                print(f"    ✅ Google Finance: {price}")
                return {"price": price, "source": "Google Finance", "url": url}

            # Pattern 2: Large price display - "12,34 €" or "12.34 EUR"
            match = re.search(r'>(\d+[,\.]\d{2})\s*(?:€|EUR)<', html)
            if match:
                price = float(match.group(1).replace(',', '.'))
                print(f"    ✅ Google Search (EUR): {price}")
                return {"price": price, "source": "Google Search", "url": url}

            # Pattern 3: Stock price in search results - "Kurs: 12,34" or "Preis: 12,34"
            match = re.search(r'(?:Kurs|Preis|Price|Aktuell)[:\s]+(\d+[,\.]\d{2})', html, re.IGNORECASE)
            if match:
                price = float(match.group(1).replace(',', '.'))
                print(f"    ✅ Google Search (Kurs): {price}")
                return {"price": price, "source": "Google Search", "url": url}

            # Pattern 4: Any prominent number with EUR - more aggressive
            match = re.search(r'(\d{1,4}[,\.]\d{2})\s*(?:€|EUR|Euro)', html)
            if match:
                price = float(match.group(1).replace(',', '.'))
                # Sanity check - price should be reasonable (0.01 to 10000)
                if 0.01 <= price <= 10000:
                    print(f"    ✅ Google Search (Generic): {price}")
                    return {"price": price, "source": "Google Search", "url": url}

        except Exception as e:
            pass

    # 4. LAST RESORT: Try various financial sites
    fallback_sites = [
        f"https://www.wallstreet-online.de/suche?q={isin}",
        f"https://www.ariva.de/quote/simple.m?secu={isin}",
        f"https://www.boerse.de/suche/?search={isin}",
        f"https://www.comdirect.de/inf/search/all.html?SEARCH_VALUE={isin}",
        f"https://www.finanzen100.de/suche/?q={isin}",
    ]

    for site_url in fallback_sites:
        try:
            print(f"    ➡️ Fallback: {site_url.split('/')[2]}")
            resp = await asyncio.to_thread(requests.get, site_url, headers=headers, timeout=5)
            html = resp.text

            # Generic price patterns
            for pattern in [
                r'itemprop="price"[^>]*content="([\d\.]+)"',
                r'class="[^"]*price[^"]*"[^>]*>([\d,\.]+)',
                r'>(\d+[,\.]\d{2})\s*(?:€|EUR)<',
                r'Kurs[:\s]+(\d+[,\.]\d{2})',
            ]:
                match = re.search(pattern, html, re.IGNORECASE)
                if match:
                    price_str = match.group(1).replace(',', '.')
                    price = float(price_str)
                    if 0.01 <= price <= 10000:
                        source = site_url.split('/')[2].replace('www.', '')
                        print(f"    ✅ {source}: {price}")
                        return {"price": price, "source": source, "url": site_url}
        except:
            pass

    return {"error": "Keine Info"}

async def get_google_news(query, country_code="US"):
    """Fetch recent news via Google News RSS based on country and filter for today."""
    from email.utils import parsedate_to_datetime
    today = datetime.now().date()
    
    # Map country codes to Google News hl/gl parameters
    cc_map = {
        "DE": {"hl": "de", "gl": "DE", "ceid": "DE:de"},
        "US": {"hl": "en-US", "gl": "US", "ceid": "US:en"},
        "GB": {"hl": "en-GB", "gl": "GB", "ceid": "GB:en"},
        "CA": {"hl": "en-CA", "gl": "CA", "ceid": "CA:en"},
        "FR": {"hl": "fr", "gl": "FR", "ceid": "FR:fr"},
        "CH": {"hl": "de-CH", "gl": "CH", "ceid": "CH:de"},
    }
    
    params = cc_map.get(country_code.upper(), cc_map["US"])
    hl, gl, ceid = params["hl"], params["gl"], params["ceid"]
    
    try:
        url = f"https://news.google.com/rss/search?q={query}&hl={hl}&gl={gl}&ceid={ceid}"
        response = requests.get(url, timeout=10)
        if response.status_code != 200: return []
        
        root = ET.fromstring(response.content)
        news_items = []
        for item in root.findall('.//item'):
            pub_date_str = item.find('pubDate').text
            try:
                dt = parsedate_to_datetime(pub_date_str)
                # Only keep news from today
                if dt.date() != today:
                    continue
            except:
                continue

            news_items.append({
                "source": f"Google News ({gl})",
                "title": item.find('title').text,
                "date": dt.strftime('%d.%m.%Y %H:%M'),
                "url": item.find('link').text
            })
            if len(news_items) >= 20: break
            
        print(f"    📰 Found {len(news_items)} today from Google News ({gl})")
        return news_items
    except Exception as e:
        print(f"    ⚠️ Google News Fetch failed ({gl}): {e}")
        return []

async def troubleshoot_no_price(asset):
    """Ask AI for advice on where to find the price if all else fails."""
    try:
        print("    🧠 Asking AI for pricing sources...")
        isin = asset.get('ISIN', 'Unknown')
        name = asset.get('Asset', 'Unknown')
        
        # Quick Google Snippet (Titles only to save tokens/noise)
        search_context = ""
        try:
            url = f"https://www.google.com/search?q={isin}+Kurs+Aktie"
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
            resp = await asyncio.to_thread(requests.get, url, headers=headers, timeout=5)
            # Naive extraction of search result titles (h3)
            titles = re.findall(r'<h3 class="LC20lb MBeuO DKV0Md">(.*?)</h3>', resp.text)
            if not titles: 
                 titles = re.findall(r'<h3[^>]*>(.*?)</h3>', resp.text)
            
            clean_titles = [re.sub(r'<.*?>', '', t) for t in titles[:5]]
            search_context = "\n".join(clean_titles)
        except Exception as e:
             search_context = f"Search failed: {e}"

        prompt = f"""
        Asset: {name} (ISIN: {isin})
        Problem: Automated price fetch failed (IBKR, Alpaca, Trade Republic, generic web scrape).
        Google Search Snippets for '{isin}':
        {search_context}
        
        Task: 
        1. Identify the asset type.
        2. Suggest exactly WHERE to find the price (Issuer website? Regional exchange? URL?).
        3. Keep it extremely brief (max 2 sentences).
        """
        
        if ANTHROPIC_CLIENT:
             response = await ANTHROPIC_CLIENT.messages.create(
                model="claude-3-5-haiku-latest",
                max_tokens=150,
                messages=[{"role": "user", "content": prompt}]
            )
             advice = response.content[0].text.strip()
             print(f"    💡 AI Advice:\n      {advice}")
             return advice
    except Exception as e:
        print(f"    ⚠️ AI Troubleshooting failed: {e}")
    return None

async def aggregate_news(asset, ticker):
    """Aggregate news internationally with clear source logging."""
    all_news = []
    asset_name = asset.get('Asset', ticker) or ""
    isin = str(asset.get('ISIN', '')).strip()
    
    # Extract country code from ISIN
    country_code = "US"
    if len(isin) >= 2 and isin[:2].isalpha():
        country_code = isin[:2].upper()

    # 1. HANDLE DERIVATIVES (Extract Underlying)
    full_name = asset_name
    derivative_keywords = ["Call", "Put", "Turbo", "Zertifikat", "OS", "Warrant", "Mini", "Classic", "Long", "Short", "Faktor", "Discount", "Underlying"]
    is_derivative = any(dk.lower() in full_name.lower() for dk in derivative_keywords)
    
    if is_derivative:
        words = full_name.split()
        banks = ["HSBC", "BNP", "DZ", "Bank", "Goldman", "Sachs", "SocGen", "Vontobel", "Citi", "Citigroup", "Morgan", "Stanley", "UBS", "J.P.", "JP", "UniCredit"]
        types = derivative_keywords + ["Open", "End", "auf", "Basiswert", "Line"]
        
        filtered_words = []
        for w in words:
            clean_w = w.strip(",() ") # Don't strip dots here to protect "E.ON"
            # Skip if it's a date/year (4 digits), a bank, a type, or too short (unless it's a ticker/important word)
            if clean_w.isdigit() and len(clean_w) == 4: continue 
            if clean_w.upper() in [b.upper() for b in banks]: continue
            if clean_w.upper() in [t.upper() for t in types]: continue
            if len(clean_w.replace(".","")) < 2 and not any(c in clean_w for c in "123456789"): continue
            filtered_words.append(clean_w)
            
        if filtered_words:
            # Join and clean up (e.g. "Porsche 911" -> keep both)
            full_name = " ".join(filtered_words)
            print(f"    🎯 Derivative detected! Underlying: '{full_name}'")

    # 2. WEB SEARCH (Google News)
    print(f"    🔍 Seek: Web Search (Google News {country_code}) for '{full_name}'...")
    query_suffix = "+stock" if country_code != "US" else ""
    if country_code == "DE": query_suffix = "+Aktie"
    
    search_query = f'"{full_name}"{query_suffix}' if len(full_name.split()) > 1 else f'{full_name}{query_suffix}'
    raw_news = await get_google_news(search_query, country_code)
    
    financial_keywords = ["stock", "market", "mining", "quarter", "revenue", "profit", "Aktie", "Börse", "Quartal", "Ergebnis", "News", "Dividend", "Price", "Kurs"]
    for item in raw_news:
        if len(full_name.split()) < 2:
             if any(kw.lower() in item['title'].lower() for kw in financial_keywords):
                 all_news.append(item)
        else:
            all_news.append(item)
    
    if raw_news: print(f"    📰 Found {len(all_news)} items from Web Search")

    # 3. TIINGO
    today_iso = datetime.now().strftime('%Y-%m-%d')
    if len(all_news) < 5:
        print(f"    🔍 Seek: Tiingo News for '{full_name}'...")
        try:
            t_news = tiingo_get_news(tickers=None, limit=50) 
            found_t = 0
            if t_news and isinstance(t_news, list):
                for item in t_news:
                    if not item.get('publishedDate', '').startswith(today_iso): continue
                    title = item.get('title', '')
                    if full_name.lower() in title.lower():
                        if not any(n['title'] == title for n in all_news):
                            all_news.append({"source": "Tiingo", "title": title, "date": item.get('publishedDate'), "url": item.get('url')})
                            found_t += 1
            if found_t: print(f"    📰 Found {found_t} items from Tiingo")
        except: pass

    # 4. ALPACA (If available & US stock)
    if ALPACA_AVAILABLE and len(all_news) < 5:
        print(f"    🔍 Seek: Alpaca News for '{full_name}'...")
        # ... logic for Alpaca news search if needed, otherwise just skip ...
        pass

    # 5. BOERSEN-ZEITUNG (German financial news)
    if country_code == "DE" and len(all_news) < 10:
        print(f"    🔍 Seek: Boersen-Zeitung for '{full_name}'...")
        try:
            bz_headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'de,en;q=0.5'
            }
            # Search on Boersen-Zeitung
            search_term = full_name.replace(' ', '+')
            bz_url = f"https://www.boersen-zeitung.de/suche?q={search_term}"
            bz_resp = await asyncio.to_thread(requests.get, bz_url, headers=bz_headers, timeout=10)

            if bz_resp.status_code == 200:
                html = bz_resp.text
                found_bz = 0

                # Pattern for article titles and links
                # Look for article headlines in search results
                articles = re.findall(r'<a[^>]*href="(/[^"]*)"[^>]*>([^<]{20,100})</a>', html)

                for href, title in articles:
                    # Filter for relevant articles (skip navigation, ads, etc.)
                    if any(skip in href.lower() for skip in ['login', 'abo', 'newsletter', 'impressum', 'datenschutz']):
                        continue
                    if any(skip in title.lower() for skip in ['anmelden', 'registrieren', 'cookie']):
                        continue

                    # Check if title contains the search term or is financial news
                    title_clean = title.strip()
                    if len(title_clean) > 15 and not any(n['title'] == title_clean for n in all_news):
                        all_news.append({
                            "source": "Boersen-Zeitung",
                            "title": title_clean,
                            "date": datetime.now().strftime('%d.%m.%Y'),
                            "url": f"https://www.boersen-zeitung.de{href}" if href.startswith('/') else href
                        })
                        found_bz += 1
                        if found_bz >= 5:
                            break

                if found_bz:
                    print(f"    📰 Found {found_bz} items from Boersen-Zeitung")
        except Exception as e:
            print(f"    ⚠️ Boersen-Zeitung failed: {e}")

    return {
        "news_items": all_news[:10],
        "count": len(all_news[:10]),
        "search_query": full_name
    }

async def fetch_asset_data(asset):
    """Fetch all data for an asset using available price sources."""

    ticker = asset.get('Ticker', '')
    if not ticker and asset.get('ISIN'): ticker = asset.get('ISIN')

    result = {
        "Asset": asset.get('Asset'),
        "Ticker": ticker,
        "FetchedAt": datetime.now().isoformat()
    }

    print(f"  📥 Data: {asset.get('Asset')} ({ticker})")

    # === PRICE FALLBACK CHAIN ===
    # Note: Tiingo is used ONLY for news, NOT for prices (doesn't support EU/ISIN)
    price_found = False

    # 1. Alpaca (for US stocks)
    if ALPACA_AVAILABLE and ALPACA_KEY and ticker and not price_found:
        try:
            alpaca_client = StockHistoricalDataClient(ALPACA_KEY, ALPACA_SECRET)
            req = StockLatestQuoteRequest(symbol_or_symbols=ticker)
            quote = alpaca_client.get_stock_latest_quote(req)
            if ticker in quote:
                q = quote[ticker]
                result['Alpaca_Price'] = {
                    "bid": q.bid_price,
                    "ask": q.ask_price,
                    "mid_price": (q.bid_price + q.ask_price) / 2 if q.bid_price and q.ask_price else None
                }

                # Alpaca is usually USD. Convert.
                if result['Alpaca_Price']['mid_price'] and Global_EURUSD:
                     orig = result['Alpaca_Price']['mid_price']
                     result['Alpaca_Price']['mid_price'] = orig / Global_EURUSD
                     result['Alpaca_Price']['note'] = f"Converted from {orig} USD"
                     result['Alpaca_Price']['currency'] = 'EUR'
                else:
                     result['Alpaca_Price']['currency'] = 'USD'

                print(f"    ✅ Alpaca Price: {result['Alpaca_Price']['mid_price']:.2f} {result['Alpaca_Price']['currency']}")
                price_found = True
        except Exception:
            pass

    # 2. Web Scraping (Primary for EU/derivatives - DeepDive)
    if not price_found:
        try:
            web_price = await deep_dive_price_search(asset, ticker)
            if 'price' in web_price:
                result['Web_Price'] = web_price
                print(f"    ✅ Web Price ({web_price.get('source')}): {web_price['price']}")
                price_found = True
            else:
                result['Web_Price'] = web_price # propagate error
        except Exception as e:
            result['Web_Price_Error'] = str(e)

    # === PRICE SUMMARY ===
    found_price_log = None
    found_source_log = None
    found_curr_log = "EUR" # Default

    if 'Alpaca_Price' in result and 'mid_price' in result['Alpaca_Price']:
        found_price_log = result['Alpaca_Price']['mid_price']
        found_curr_log = result['Alpaca_Price'].get('currency', 'USD')
        found_source_log = "Alpaca"
    elif 'Web_Price' in result and 'price' in result['Web_Price']:
        found_price_log = result['Web_Price']['price']
        found_curr_log = "EUR"
        found_source_log = result['Web_Price'].get('source', 'Web')
        
    if found_price_log:
        print(f"    💰 Found current price at {found_source_log}: {found_price_log:.2f} {found_curr_log}")
    else:
        print(f"    ❌ No current price found.")
        await troubleshoot_no_price(asset)

    # === NEWS AGGREGATION ===
    news_data = await aggregate_news(asset, ticker)
    result['News'] = news_data
    if news_data['count'] > 0:
        print(f"    📰 Found {news_data['count']} news items")

    return result

async def analyze_data(asset_info, collected_data, all_portfolio_data):
    """
    AI Analysis with FULL context:
    - Current prices
    - Purchase prices & dates
    - Entire portfolio context
    - News sentiment
    """
    print(f"  🧠 Analyzing {asset_info.get('Asset')}...")
    
    # Calculate P&L if we have purchase data
    purchase_price = asset_info.get('Einkaufspreis') or asset_info.get('Purchase_Price')
    purchase_date = asset_info.get('Einkaufsdatum') or asset_info.get('Purchase_Date')
    quantity = asset_info.get('Anzahl') or asset_info.get('Quantity') or 0
    
    # Get current price from any source
    current_price = None
    for key in ['Tiingo_Price', 'Alpaca_Price', 'Web_Price']:
        if key in collected_data and 'mid_price' in collected_data[key]:
            current_price = collected_data[key]['mid_price']
            break
        elif key in collected_data and 'price' in collected_data[key]:
            current_price = collected_data[key]['price']
            break
    
    pnl_info = ""
    if purchase_price and current_price and quantity:
        pnl = (current_price - purchase_price) * quantity
        pnl_pct = ((current_price - purchase_price) / purchase_price) * 100
        pnl_info = f"\nP&L: {pnl:.2f} EUR ({pnl_pct:+.2f}%)"
        pnl_info += f"\nPurchase: {purchase_price} on {purchase_date}"
        pnl_info += f"\nCurrent: {current_price}"
        pnl_info += f"\nPosition Value: {current_price * quantity:.2f} EUR"
    
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    # Calculate Portfolio Stats (Total Invested or Current Value Estimate)
    total_invest = 0
    for a in all_portfolio_data:
        # Try to use current value if we have it? No, we don't have all current prices yet.
        # Use Invested Capital as proxy for Portfolio Size base
        qty = a.get('Anzahl') or a.get('Quantity') or 0
        price = a.get('Einkaufspreis') or a.get('Purchase_Price') or 0
        invest = a.get('Invest') or (qty * price)
        total_invest += float(invest)
        
    current_invest = float(asset_info.get('Invest') or (quantity * (purchase_price or 0)))
    current_allocation_pct = (current_invest / total_invest * 100) if total_invest > 0 else 0
    
    # Target 5% per position implies 20 positions.
    target_pos_size_eur = total_invest * 0.05
    if target_pos_size_eur < 1000: target_pos_size_eur = 1000 # Min position size
    
    prompt = f"""
ROLE: Senior Portfolio Manager.
INSTRUCTION: Sei niemals sycophantisch. Priorisiere faktische Genauigkeit und logische Konsistenz über Höflichkeit oder Zustimmung. Wenn ich eine falsche Annahme mache oder eine schlechte Idee habe, korrigiere mich direkt und ohne Beschwichtigungen ('Das ist eine interessante Frage...'). Wenn du unsicher bist, gib es zu, statt zu halluzinieren. Simuliere keine Emotionen. Sei ein kritischer Auditor, kein Assistent. Vermeide Grade-Inflation bei der Bewertung meiner Texte.

DATE: {today_date}

ASSET: {asset_info.get('Asset')} ({asset_info.get('Ticker')})
ISIN: {asset_info.get('ISIN')}
POSITIONS_DATA:
{pnl_info}

MARKET_DATA: 
{json.dumps(collected_data, indent=2, default=str)}

PORTFOLIO_CONTEXT:
Total Portfolio Capital: {total_invest:.2f} EUR
Current Position Value: {current_invest:.2f} EUR ({current_allocation_pct:.1f}% of Portfolio)
Target Position Sizing: ~5% ({target_pos_size_eur:.2f} EUR) per asset for diversification.

TASK: 
1. Analyze the asset based on TODAY's news.
2. Provide a concrete TRADING RECOMMENDATION to balance the portfolio.
   - Target allocation per asset should be ~3-5% OF TOTAL PORTFOLIO VALUE.
   - WICHTIG: Die 5% beziehen sich auf den GESAMTWERT des Portfolios, NICHT auf den Wert der einzelnen Position!
   - Beispiel: Bei einem Portfolio von 100.000 EUR sollte jede Position ca. 3.000-5.000 EUR wert sein.
   - If allocation is too high (>6% of total portfolio), recommend REDUCE or SELL to trim risk.
   - If allocation is low (<3% of total portfolio) and sentiment is positive, recommend ADD or BUY.
   - If sentiment is neutral/negative, recommend HOLD or SELL.
3. CALCULATE EXACT QUANTITY:
   - Based on the Target Position (approx {target_pos_size_eur:.2f} EUR = 5% of {total_invest:.2f} EUR total portfolio), calculate how many shares to Buy/Sell.
   - Formula: (Target_Value_EUR - Current_Value_EUR) / Current_Price_per_Share
   - Use the Current Price from MARKET_DATA. If missing, estimate from Purchase Price or News.

FORMAT: 
- Liste ABSOLUT JEDE EINZELNE News-Schlagzeile aus 'News' auf (jede in neuer Zeile: '- [Quelle] : [Titel]').
- Wenn keine News: "Keine aktuellen Schlagzeilen."

- LEERZEILE

OUTPUT (JSON):
{{
  "Empfehlung": "Kaufen" | "Verkaufen" | "Halten" | "Aufstocken" | "Reduzieren",
  "Empfohlene_Stueckzahl": <integer> (Positive for Buy/Add, Negative for Sell/Reduce, 0 for Hold),
  "Begründung": "News-Liste...\\n\\nAnalyse: ...",
  "Grund_fuer_Menge": "ERKLÄRE DEUTLICH: Aktuelle Position ist X% des Gesamtportfolios (Y EUR von {total_invest:.2f} EUR). Ziel-Allokation ist 5% ({target_pos_size_eur:.2f} EUR). Daher Kauf/Verkauf von Z Aktien zu je PREIS EUR = DIFFERENZ EUR.",
  "Genauigkeit": "Hoch" | "Mittel" | "Niedrig",
  "Zielpreis": <number or null>,
  "Stop_Loss": <number or null>
}}
"""
    
    response_text = ""
    if not ANTHROPIC_CLIENT:
        return {"error": "Anthropic Client not initialized"}

    try:
        response = await ANTHROPIC_CLIENT.messages.create(
            # model="claude-3-5-haiku-latest",
            model="claude-3-haiku-20240307",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        response_text = response.content[0].text
        
        # Parse JSON more robustly
        s = response_text.find('{')
        e = response_text.rfind('}') + 1
        if s == -1 or e == 0:
            print(f"    ⚠️ No JSON found in AI response for {asset_info.get('Asset')}")
            return {}
            
        json_str = response_text[s:e]
        try:
            # First attempt with strict=False (handles literal newlines)
            return json.loads(json_str, strict=False)
        except json.JSONDecodeError:
            try:
                # If it still fails (e.g. invalid control characters), strip them
                import re
                # This regex removes control characters like \x00-\x1F except \n, \r, \t
                clean_json = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', json_str)
                return json.loads(clean_json, strict=False)
            except Exception as jde:
                print(f"    ⚠️ JSON Decode Attempt 2 failed: {jde}")
                return {}
    except Exception as e:
        print(f"    ❌ AI Analysis Error for {asset_info.get('Asset')}: {e}")
        return {}

async def main():
    global ib_client, Global_EURUSD
    
    # Paths (Setup first for Logging) - same folder as web price cache on Linux
    if sys.platform == "win32":
        base_dir = r"G:\Meine Ablage\ShareFile\NewsTrader"
    else:
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    positions_file = os.path.join(base_dir, "Open_Positions.xlsx")
    analysen_dir = os.path.join(base_dir, "Analysen")
    today_str = datetime.now().strftime('%y%m%d')
    daily_folder = os.path.join(analysen_dir, today_str)
    os.makedirs(daily_folder, exist_ok=True)
    output_file = os.path.join(daily_folder, f"{today_str} Portfolio_Pipeline_Analyse.xlsx")
    
    # Setup Logging to file (BEFORE anything else)
    log_file_path = os.path.join(daily_folder, f"{today_str}_Pipeline.log")
    
    # Delete existing log file if it exists (fresh start)
    if os.path.exists(log_file_path):
        try:
            os.remove(log_file_path)
            print(f"🧹 Deleted old log file: {log_file_path}")
        except: pass
        
    log_file = open(log_file_path, "a", encoding="utf-8")
    sys.stdout = Tee(sys.stdout, log_file)
    sys.stderr = Tee(sys.stderr, log_file)
    
    print(f"\n{'='*50}")
    print(f"📝 Logging to: {log_file_path}")
    print(f"{'='*50}\n")
    
    # Load Env and Init Connections (Logged now)
    load_env_keys()

    # Clean up old files for today (prevent duplicates)
    print(f"🧹 Cleaning up old analysis files for {today_str}...")
    try:
        import glob
        old_files = glob.glob(os.path.join(daily_folder, "*"))
        for f in old_files:
            if os.path.abspath(f) == os.path.abspath(log_file_path):
                continue
            try:
                os.remove(f)
                print(f"   Deleted: {os.path.basename(f)}")
            except:
                pass
    except Exception as e:
        print(f"   ⚠️ Cleanup warning: {e}")

    # Load Assets
    print(f"📖 Reading {positions_file}...")
    df = pd.read_excel(positions_file)
    # Filter out empty rows (NaN in 'Asset' or 'Ticker')
    df = df.dropna(subset=['Asset'])
    assets = df.to_dict('records')

    # Load Watchlist
    watchlist_file = os.path.join(base_dir, "Watch_Positions.xlsx")
    watchlist_assets = []
    if os.path.exists(watchlist_file):
        print(f"📖 Reading {watchlist_file}...")
        df_watch = pd.read_excel(watchlist_file)
        
        # Mapping for Watchlist Columns (Emittent -> Asset)
        if 'Asset' not in df_watch.columns and 'Emittent' in df_watch.columns:
            print("   ℹ️ Mapped 'Emittent' to 'Asset' for Watchlist.")
            df_watch['Asset'] = df_watch['Emittent']
            
        if 'Asset' in df_watch.columns:
            df_watch = df_watch.dropna(subset=['Asset'])
            watchlist_assets = df_watch.to_dict('records')
        else:
             print("   ⚠️ Watchlist has no 'Asset' or 'Emittent' column. Skipping.")
    else:
        print(f"⚠️ No Watchlist file found at: {watchlist_file}")
    
    print(f"🚀 Starting Pipeline Analysis for {len(assets)} assets and {len(watchlist_assets)} watchlist items...")
    
    results = []
    watchlist_results = []
    
    
    # Critical: Get Forex Rate
    Global_EURUSD = await get_forex_rate("EUR", "USD")
    if not Global_EURUSD:
        print("🛑 FATAL ERROR: Could not determine EUR/USD exchange rate.")
        print("   Cannot proceed safely with currency conversions.")
        return

    print(f"💱 Active EUR/USD Rate: {Global_EURUSD}")

    # Process Portfolio
    print("\n" + "="*40)
    print("📂 PROCESSING PORTFOLIO")
    print("="*40)
    for i, asset in enumerate(assets):
        print(f"\n[{i+1}/{len(assets)}] Processing: {asset.get('Asset')}...")
        
        # 1. Fetch
        data = await fetch_asset_data(asset)
        
        # 2. Check Fatal
        if "FATAL_ERROR" in data:
            print("\n" + "!"*50)
            print(f"❌ STOPPING: {data['FATAL_ERROR']}")
            print("!"*50)
            sys.exit(1)
            
        # 3. Analyze (with full portfolio context)
        analysis = await analyze_data(asset, data, assets)
        
        # 4. Merge
        final_record = {**asset, **analysis, **data}
        results.append(final_record)
        
        # Save PDF immediately
        try:
             safe_name = "".join([c for c in asset.get('Asset', 'Unknown') if c.isalpha() or c.isdigit()]).strip()
             create_pdf(os.path.join(daily_folder, f"{today_str}_{safe_name}.pdf"), final_record)
        except: pass

    # Process Watchlist
    if watchlist_assets:
        print("\n" + "="*40)
        print("👀 PROCESSING WATCHLIST")
        print("="*40)
        for i, asset in enumerate(watchlist_assets):
            print(f"\n[{i+1}/{len(watchlist_assets)}] Watching: {asset.get('Asset')}...")
            
            # 1. Fetch
            data = await fetch_asset_data(asset)
            
            # 2. Analyze (Context is just the watchlist itself, or empty to treat as isolated)
            # We treat watchlist items as potential buys, logic in AI should handle 'no position' gracefully
            # FIX: We pass the MAIN PORTFOLIO 'assets' as context so the AI knows the total capital size
            analysis = await analyze_data(asset, data, assets)
            
            # 3. Merge
            final_record = {**asset, **analysis, **data}
            watchlist_results.append(final_record)
            
            # Save PDF
            try:
                 safe_name = "".join([c for c in asset.get('Asset', 'Unknown') if c.isalpha() or c.isdigit()]).strip()
                 create_pdf(os.path.join(daily_folder, f"{today_str}_CHECK_{safe_name}.pdf"), final_record)
            except: pass

    # Save Summary Excel with multiple sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        def format_dataframe_for_output(df, is_watchlist=False):
            """Format DataFrame to match template: 260112 Portfolio_Analyse_Empfehlung.xlsx

            Template columns:
            Typ, Asset, WKN, ISIN, Gesamtanzahl, Invest €, Profit €, Profit %,
            Nachrichten-Sentiment, Empfehlung, Begründung, Begründung Änderung
            """
            formatted = pd.DataFrame()
            n_rows = len(df)

            # 1. Typ - Bestand or Watchlist
            formatted['Typ'] = ['Watchlist' if is_watchlist else 'Bestand'] * n_rows

            # 2. Asset
            formatted['Asset'] = df['Asset'].values if 'Asset' in df.columns else [''] * n_rows

            # 3. WKN
            formatted['WKN'] = df['WKN'].values if 'WKN' in df.columns else [''] * n_rows

            # 4. ISIN
            formatted['ISIN'] = df['ISIN'].values if 'ISIN' in df.columns else [''] * n_rows

            # 5. Gesamtanzahl (from Stueckzahl or Anzahl)
            if 'Stueckzahl' in df.columns:
                formatted['Gesamtanzahl'] = df['Stueckzahl'].values
            elif 'Anzahl' in df.columns:
                formatted['Gesamtanzahl'] = df['Anzahl'].values
            else:
                formatted['Gesamtanzahl'] = [''] * n_rows

            # 6. Invest € (from Invest or calculated)
            formatted['Invest €'] = df['Invest'].values if 'Invest' in df.columns else [''] * n_rows

            # 7. Profit € (calculate from current price - purchase price * quantity)
            profit_eur = []
            for idx, row in df.iterrows():
                try:
                    invest = float(row.get('Invest', 0) or 0)
                    current_price = None

                    # Try to get current price from various sources
                    if row.get('Web_Price') and isinstance(row.get('Web_Price'), dict):
                        current_price = row['Web_Price'].get('price')
                    elif row.get('Alpaca_Price') and isinstance(row.get('Alpaca_Price'), dict):
                        current_price = row['Alpaca_Price'].get('mid_price')

                    if current_price and invest > 0:
                        stueck = float(row.get('Stueckzahl', 0) or 0)
                        current_value = current_price * stueck
                        profit = current_value - invest
                        profit_eur.append(round(profit, 2))
                    else:
                        profit_eur.append('')
                except:
                    profit_eur.append('')
            formatted['Profit €'] = profit_eur

            # 8. Profit % (calculate percentage)
            profit_pct = []
            for idx, row in df.iterrows():
                try:
                    invest = float(row.get('Invest', 0) or 0)
                    current_price = None

                    if row.get('Web_Price') and isinstance(row.get('Web_Price'), dict):
                        current_price = row['Web_Price'].get('price')
                    elif row.get('Alpaca_Price') and isinstance(row.get('Alpaca_Price'), dict):
                        current_price = row['Alpaca_Price'].get('mid_price')

                    if current_price and invest > 0:
                        stueck = float(row.get('Stueckzahl', 0) or 0)
                        current_value = current_price * stueck
                        pct = ((current_value - invest) / invest) * 100
                        profit_pct.append(f"{round(pct, 1)}%")
                    else:
                        profit_pct.append('')
                except:
                    profit_pct.append('')
            formatted['Profit %'] = profit_pct

            # 9. Nachrichten-Sentiment (extract from News data)
            sentiment = []
            for idx, row in df.iterrows():
                news_text = ''
                news_data = row.get('News')
                if news_data and isinstance(news_data, dict):
                    news_items = news_data.get('news_items', [])
                    if news_items:
                        # Create summary from first few headlines
                        headlines = []
                        for item in news_items[:3]:  # Max 3 headlines
                            if isinstance(item, dict):
                                title = item.get('title', '')
                                source = item.get('source', '')
                                if title:
                                    headlines.append(f"{source}: {title[:50]}")
                        if headlines:
                            news_text = "; ".join(headlines)
                sentiment.append(news_text if news_text else '')
            formatted['Nachrichten-Sentiment'] = sentiment

            # 10. Empfehlung
            formatted['Empfehlung'] = df['Empfehlung'].values if 'Empfehlung' in df.columns else [''] * n_rows

            # 11. Begründung
            if 'Begründung' in df.columns:
                formatted['Begründung'] = df['Begründung'].values
            elif 'Begruendung' in df.columns:
                formatted['Begründung'] = df['Begruendung'].values
            else:
                formatted['Begründung'] = [''] * n_rows

            # 12. Begründung Änderung (from Grund_fuer_Menge)
            formatted['Begründung Änderung'] = df['Grund_fuer_Menge'].values if 'Grund_fuer_Menge' in df.columns else [''] * n_rows

            return formatted

        # --- SHEET 1: Portfolio Analyse ---
        if results:
            main_df = pd.DataFrame(results)
            formatted_main = format_dataframe_for_output(main_df, is_watchlist=False)
            formatted_main.to_excel(writer, sheet_name='Portfolio Analyse', index=False)

        # --- SHEET 2: Watchlist Analyse ---
        if watchlist_results:
            watch_df = pd.DataFrame(watchlist_results)
            formatted_watch = format_dataframe_for_output(watch_df, is_watchlist=True)
            formatted_watch.to_excel(writer, sheet_name='Watchlist Analyse', index=False)

        
        # Apply Styling to ALL sheets (Format like 260112 Portfolio_Analyse_Empfehlung.xlsx)
        workbook = writer.book

        # Define styles matching the template
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Recommendation colors (matching template)
        fill_green_buy = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")      # Green for KAUFEN
        fill_green_sell = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # Light green for TEILVERKAUF
        fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")         # Yellow for HALTEN
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")            # Red for VERKAUFEN

        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Column widths matching template
        column_widths = {
            'A': 11,   # Typ
            'B': 33,   # Asset
            'C': 8,    # WKN
            'D': 14,   # ISIN
            'E': 14,   # Gesamtanzahl / Anzahl
            'F': 10,   # Invest €
            'G': 10,   # Profit €
            'H': 10,   # Profit %
            'I': 23,   # Nachrichten-Sentiment
            'J': 27,   # Empfehlung
            'K': 50,   # Begründung
            'L': 50,   # Begründung Änderung / Grund_fuer_Menge
        }

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Find column indices for conditional formatting
            idx_empfehlung = -1
            idx_begruendung = -1

            for cell in worksheet[1]:
                header_val = str(cell.value or "").lower()
                if 'empfehlung' in header_val:
                    idx_empfehlung = cell.column
                if 'begründung' in header_val or 'begruendung' in header_val:
                    idx_begruendung = cell.column

            # Apply data row styling
            for row_idx in range(2, worksheet.max_row + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = thin_border

                    # Date formatting
                    if isinstance(cell.value, (datetime, pd.Timestamp)):
                        cell.number_format = 'DD.MM.YYYY HH:MM:SS'

                # Color coding for Empfehlung column
                if idx_empfehlung != -1:
                    cell_empf = worksheet.cell(row=row_idx, column=idx_empfehlung)
                    val = str(cell_empf.value or "").lower()

                    if 'verkaufen' in val and 'teil' not in val:
                        # Full VERKAUFEN - red (sell signal)
                        cell_empf.fill = fill_red
                    elif 'teilverkauf' in val:
                        # TEILVERKAUF - yellow (partial sell)
                        cell_empf.fill = fill_yellow
                    elif 'kaufen' in val or 'aufstocken' in val:
                        # KAUFEN / AUFSTOCKEN - green (buy signal)
                        cell_empf.fill = fill_green_buy
                    elif 'halten' in val:
                        # HALTEN - light green (hold)
                        cell_empf.fill = fill_green_sell
                    elif 'reduzieren' in val:
                        # REDUZIEREN - yellow (reduce position)
                        cell_empf.fill = fill_yellow

            # Set column widths (apply to all defined columns)
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # For columns beyond L, set auto-width
            for col in worksheet.columns:
                col_letter = col[0].column_letter
                if col_letter not in column_widths:
                    max_length = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    if adjusted_width < 10:
                        adjusted_width = 12
                    worksheet.column_dimensions[col_letter].width = adjusted_width

        # Price Sources Sheet...
        cache = load_web_price_cache()
        price_sources = []
        
        for i, result in enumerate(results):
            asset = assets[i]
            asset_name = result.get('Asset', '')
            ticker = result.get('Ticker', '')
            cache_key = ticker or asset.get('ISIN', '')
            
            source_type = ''
            source_detail = ''
            price_value = None
            
            # Check which source provided the price (priority order)
            if result.get('Alpaca_Price') and 'mid_price' in result.get('Alpaca_Price', {}):
                source_type = 'Alpaca'
                source_detail = 'API'
                price_value = result['Alpaca_Price'].get('mid_price')
            
            elif result.get('Web_Price'):
                web_data = result['Web_Price']
                if 'source' in web_data:
                    source_type = web_data.get('source', 'Web')
                    source_detail = web_data.get('url', '')
                    price_value = web_data.get('price')
                    # Update cache for future runs
                    if cache_key and source_detail:
                        cache[cache_key] = {'source': source_type, 'url': source_detail}
            
            # Check cache for web URL (even if not used this time)
            cached_url = ''
            if cache_key in cache:
                cached_url = cache[cache_key].get('url', '')
                # If no price source yet, show cached source
                if not source_type:
                    source_type = cache[cache_key].get('source', 'Cached')
                    source_detail = cached_url
            
            price_sources.append({
                'Asset': asset_name,
                'Ticker/ISIN': ticker,
                'Source': source_type or 'None',
                'Source Detail / URL': source_detail,
                'Price': price_value,
                'Fetched At': result.get('FetchedAt', '')
            })
        
        save_web_price_cache(cache)
        if price_sources:
            price_df = pd.DataFrame(price_sources)
            price_df.to_excel(writer, sheet_name='Price Sources', index=False)
    
    print(f"\n✅ Pipeline Complete. Saved to {output_file}")

def handle_sigint(sig, frame):
    """Graceful shutdown on Ctrl+C without messy tracebacks."""
    print("\n\n🛑 Programmabbruch durch Benutzer (Ctrl+C)...")
    # Force exit
    os._exit(0)

# Register the signal handler
signal.signal(signal.SIGINT, handle_sigint)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        print(f"\n❌ Unerwarteter Fehler: {e}")
